import pandas as pd
import os
#from shutil import copyfile
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog #, messagebox

def process_rna_aliquots(input_file, output_file, extract_type, status_label):
    status_label.config(text="Processing aliquot data...")
    # read the input Excel file
    df = pd.read_excel(input_file, engine="openpyxl")

    processed_data = []

    for _, row in df.iterrows():
        tissue_id = row["GL Sample ID"]

        # handle ALQ0 (remaining uL)
        if pd.notna(row.get("ALQ0 (remaining uL)")):
            processed_data.append({
                "Subject Barcode" : row["Barcode"],
                "GL Sample ID": tissue_id,
                "Analyzed Sample Weight (mg)": row["Analyzed Sample Weight (mg)"],
                "RNA Conc. (ng/ul) BR": row["RNA Conc. (ng/ul) BR"],
                "RNA Volume (ul)": row["RNA Volume (ul)"],
                "RNA Yield (ug)": row["RNA Yield (ug)"],
                "RIN (Bioanalyzer, Nano)": row["RIN (Bioanalyzer, Nano)"],
                "DV200 (%)": row["DV200 (%)"],
                "Date Processed": row["Date Processed"],
                "Water Added ALQ1": row["uL H2O to ALQ 1"],
                "Extract ID": f"{tissue_id}_{extract_type}_ALQ0",
                "Aliquot Volume (uL)": row["ALQ0 (remaining uL)"],
                "Box ID": row["Sample box w/ data matrix barcode"],
                "Aliquot Box Location": row["ALQ0 in Box"]
            })

        # handle ALQ1 (uL to ALQ1)
        if pd.notna(row.get("ALQ 1 (uL = 1.5 ug)")):
            processed_data.append({
                "Subject Barcode" : row["Barcode"],
                "GL Sample ID": tissue_id,
                "Analyzed Sample Weight (mg)": row["Analyzed Sample Weight (mg)"],
                "RNA Conc. (ng/ul) BR": row["RNA Conc. (ng/ul) BR"],
                "RNA Volume (ul)": row["RNA Volume (ul)"],
                "RNA Yield (ug)": row["RNA Yield (ug)"],
                "RIN (Bioanalyzer, Nano)": row["RIN (Bioanalyzer, Nano)"],
                "DV200 (%)": row["DV200 (%)"],
                "Date Processed": row["Date Processed"],
                "Water Added ALQ1": row["uL H2O to ALQ 1"],
                "Extract ID": f"{tissue_id}_{extract_type}_ALQ1",
                "Aliquot Volume (uL)": row["ALQ 1 (uL = 1.5 ug)"],
                "Box ID": row["Sample box w/ data matrix barcode"],
                "Aliquot Box Location": row["ALQ1 in Box"]
            })

        # handle ALQ2-4 (uL = 1.5 ug)
        if pd.notna(row.get("ALQ2-4 (uL = 1.5 ug)")):
            total_volume = row["ALQ2-4 (uL = 1.5 ug)"]
            for i in range(2, 5):  # Aliquots 2, 3, and 4
                aliquot_box_col = f"ALQ{i} in Box"
                if pd.notna(row.get(aliquot_box_col)):
                    processed_data.append({
                        "Subject Barcode" : row["Barcode"],
                        "GL Sample ID": tissue_id,
                        "Analyzed Sample Weight (mg)": row["Analyzed Sample Weight (mg)"],
                        "RNA Conc. (ng/ul) BR": row["RNA Conc. (ng/ul) BR"],
                        "RNA Volume (ul)": row["RNA Volume (ul)"],
                        "RNA Yield (ug)": row["RNA Yield (ug)"],
                        "RIN (Bioanalyzer, Nano)": row["RIN (Bioanalyzer, Nano)"],
                        "DV200 (%)": row["DV200 (%)"],
                        "Date Processed": row["Date Processed"],
                        "Water Added ALQ1": row["uL H2O to ALQ 1"],
                        "Extract ID": f"{tissue_id}_{extract_type}_ALQ{i}",
                        "Aliquot Volume (uL)": total_volume / 3,  # Divide volume equally
                        "Box ID": row["Sample box w/ data matrix barcode"],
                        "Aliquot Box Location": row[aliquot_box_col]
                    })

    output_df = pd.DataFrame(processed_data)

    # write the processed data to the output Excel file
    output_df.to_excel(output_file, index=False, engine="openpyxl")
    status_label.config(text="Processing complete. Data saved to output file.")
    print(f"Processed data has been written to {output_file}")

def populate_template(template_file, output_file, box_samples):

    wb = load_workbook(template_file)
    ws = wb.active

    # build a mapping of template locations to row indices
    location_to_row = {str(ws.cell(row=row, column=1).value): row for row in range(1, ws.max_row + 1)}

    # populate the template
    for _, row in box_samples.iterrows():
        location = row["Aliquot Box Location"]  # e.g. "A1", "B7"
        sample_id = row["GL Sample ID"]
        extract_id = row["Extract ID"]

        if location in location_to_row:
            row_idx = location_to_row[location]
            ws.cell(row=row_idx, column=2, value=extract_id)  # Column B: Unique Tube Barcode
            ws.cell(row=row_idx, column=4, value=row["Box ID"])  # Column D: Box ID
            ws.cell(row=row_idx, column=5, value=sample_id)  # Column E: GL Sample ID
        else:
            print(f"Warning: Location {location} not found in the template.")

    wb.save(output_file)
    #print(f"Generated template for box at {output_file}")

def generate_box_templates(output_file, template_file):
    df = pd.read_excel(output_file, engine='openpyxl')

    # Group samples by box name
    grouped_boxes = df.groupby("Box ID")

    output_csvs = []

    for box_name, box_samples in grouped_boxes:
        output_xlsx = f"box_{box_name}.xlsx"
        output_csv = f"box_{box_name}.csv"

        # populate the template and save as Excel
        populate_template(template_file, output_xlsx, box_samples)

        wb = load_workbook(output_xlsx)
        ws = wb.active
        csv_data = [[cell.value for cell in row] for row in ws.iter_rows()]
        pd.DataFrame(csv_data).to_csv(output_csv, index=False, header=False)

        output_csvs.append(output_csv)
        #print(f"Converted {output_xlsx} to {output_csv}")

    return output_csvs


def main_gui():
    root = tk.Tk()
    root.title("RNA/DNA Aliquot Processor")

    # main window
    root.geometry("500x300")

    extract_type = tk.StringVar(value="RNA")
    input_file = None
    output_dir = None

    #template_file = "Assets/81_sample_template.xlsx"

    instructions_frame = tk.Frame(root)
    instructions_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

    tk.Label(instructions_frame, text="Instructions", font=("Helvetica", 16, "bold")).pack(anchor="n")
    steps = [
        "1. Prepare sample sheet",
        "2. Download SLIMS sample metadata",
        "3. Associate SLIMS barcode to each sample",
        "4. Select files",
        "5. Run program",
        "6. Upload .csv box files"
    ]
    for step in steps:
        tk.Label(instructions_frame, text=step, anchor="w", justify="left", wraplength=300).pack(anchor="w", pady=5)

    # file selection and controls
    controls_frame = tk.Frame(root)
    controls_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

    # extract type selection
    tk.Label(controls_frame, text="Select Extract Type:").pack(pady=5)
    tk.Radiobutton(controls_frame, text="RNA", variable=extract_type, value="RNA").pack()
    tk.Radiobutton(controls_frame, text="DNA", variable=extract_type, value="DNA").pack()

    # file selection buttons
    def select_input_file():
        nonlocal input_file
        input_file = filedialog.askopenfilename(title="Select Input File", filetypes=[("Excel files", "*.xlsx")])
        status_label.config(text=f"Selected Input File: {input_file}")

    def select_output_dir():
        nonlocal output_dir
        output_dir = filedialog.askdirectory(title="Select Output Directory")
        status_label.config(text=f"Selected Output Directory: {output_dir}")

    tk.Button(controls_frame, text="Select Input File", command=select_input_file).pack(pady=5)
    tk.Button(controls_frame, text="Select Output Directory", command=select_output_dir).pack(pady=5)

    # run processing command
    def run_processing():
        if not input_file or not output_dir:
            status_label.config(text="Error: Please select all required files and directories.")
            return

        output_file = os.path.join(output_dir, "output_extract_list_test_aq22.xlsx")
        process_rna_aliquots(input_file, output_file, extract_type.get(), status_label)
        output_csvs = generate_box_templates(output_file, template_file="assets/81_sample_template.xlsx")

        print("Generated CSV files:")
        for csv_file in output_csvs:
            print(csv_file)

    tk.Button(controls_frame, text="Run Processing", command=run_processing).pack(pady=20)

    # status label at the bottom
    status_label = tk.Label(root, text="Ready!", wraplength=400, justify="center", anchor="s")
    status_label.pack(side="bottom", fill="x", pady=10)

    root.mainloop()

if __name__ == "__main__":
    main_gui()
